from openpyxl import Workbook, load_workbook

from openpyxl.cell import get_column_letter

input_workbook = load_workbook('input.xlsx')
input_worksheet = input_workbook.worksheets[0]

output_workbook = Workbook()
output_worksheet = output_workbook.get_active_sheet()


# start with row #1 and increment it in a loop until no data is present
row_number = 1


while input_worksheet.cell('A%d' % row_number).value:
    name = input_worksheet.cell('A%d' % row_number).value
    phones = input_worksheet.cell('B%d' % row_number).value
    phones_list = phones.split(',')

    print name, phones_list

    output_worksheet.cell('A%d' % row_number).value = name

    for i, phone in enumerate(phones_list):
        column_letter = get_column_letter(i + 2)
        cell_name = '%s%d' % (
                column_letter,
                row_number
                )
        output_worksheet.cell(cell_name).value = phone.strip()

    row_number = row_number + 1


output_workbook.save(filename='output.xlsx')
